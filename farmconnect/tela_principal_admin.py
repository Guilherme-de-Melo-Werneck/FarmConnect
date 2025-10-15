import flet as ft
from database import listar_medicamentos, medicamentos_mais_solicitados, editar_medicamento, adicionar_medicamento, listar_categorias, listar_fabricantes, desativar_medicamento, listar_farmacias, listar_usuarios, listar_medicamentos, registrar_usuario, listar_reagendamentos, aprovar_usuario, recusar_usuario, listar_agendamentos, adicionar_agendamento, reativar_medicamento, adicionar_estoque, aprovar_agendamento, cancelar_agendamento, confirmar_retirada_medicamento
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sqlite3
from datetime import datetime
from collections import Counter
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.lib.colors import HexColor, lightgrey, black
from collections import Counter, defaultdict


class TelaAdminDashboard:
    def __init__(self, page: ft.Page):
        self.page = page
        self.page_settings()
        self.sidebar_open = True
        self.current_view = ft.Column(expand=True)
        self.load_dashboard()  
        self.tabela_agendamentos_ref = ft.Ref[ft.DataTable]()
        self.tabela_pacientes_ref = ft.Ref[ft.DataTable]()
        self.tabela_medicamentos_ref = ft.Ref[ft.DataTable]()

    def page_settings(self):
        self.page.title = "FarmConnect - Painel Admin"
        self.page.bgcolor = "#EFF6FF"
        self.page.padding = 0
        self.page.scroll = ft.ScrollMode.ADAPTIVE
        self.page.theme_mode = ft.ThemeMode.LIGHT

    def toggle_sidebar(self, e):
        self.sidebar_open = not self.sidebar_open
        self.page.update()

    def confirmar_retirada(self, agendamento_id):
        sucesso = confirmar_retirada_medicamento(agendamento_id)

        if sucesso:
            self.page.snack_bar.content.value = "✅ Retirada confirmada com sucesso!"
            self.page.snack_bar.bgcolor = ft.Colors.GREEN_500
        else:
            self.page.snack_bar.content.value = "⚠️ Retirada já registrada anteriormente."
            self.page.snack_bar.bgcolor = ft.Colors.AMBER_500

        self.page.snack_bar.open = True
        self.page.update()

        # 🔁 Atualiza a lista/tabela após o clique
        self.load_agendamentos()

    def confirmar_agendamento(self, agendamento_id):
        aprovar_agendamento(agendamento_id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Agendamento confirmado!"), bgcolor="green")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_agendamentos()

    def cancelar_agendamento(self, agendamento_id):
        cancelar_agendamento(agendamento_id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Agendamento cancelado!"), bgcolor="red")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_agendamentos()

    def count_agendamentos_do_dia(self):
        hoje_str = datetime.now().strftime("%Y-%m-%d")
        agendamentos = listar_agendamentos()
        total_hoje = 0
        for a in agendamentos:
            # a[8] = data_criacao, ex: "2025-08-10 14:22:31"
            if a[8] and a[8].startswith(hoje_str):
                total_hoje += 1
        return total_hoje
    
    def count_agendamentos_pendentes(self):
        agendamentos = listar_agendamentos()
        total_pendentes = sum(1 for a in agendamentos if a[7] == "Pendente")
        return total_pendentes
    
    def desativar_medicamento(self, id):
        desativar_medicamento(id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Medicamento desativado."), bgcolor="red")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_medicamentos()

    def reativar_medicamento(self, id):
        reativar_medicamento(id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Medicamento reativado."), bgcolor="green")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_medicamentos()
    
    def card_estoque_medicamentos(self):
        medicamentos = listar_medicamentos(include_inativos=False)
        total_estoque = sum([m[5] or 0 for m in medicamentos]) # m[5] = estoque
        abaixo_limite = [(m[1], m[5] or 0) for m in medicamentos if (m[5] or 0) < 5]  # m[1] = nome

        if not medicamentos: # Nenhum medicamento ativo no banco
            return ft.Container(
                col={"sm": 12, "md": 4},
                bgcolor="#FFFFFF",
                border_radius=12,
                shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                padding=20,
                content=ft.Column([
                    ft.Text("Estoque de Medicamentos", size=16, weight="bold", color="#111827"),
                    ft.Text("Nenhum medicamento ativo.", size=14, color=ft.Colors.RED_700),
                ], spacing=10)
            )

        lista_criticos = ft.Column(
            controls=[
                ft.Text(f"- {nome} ({qtd} un.)", size=12, color=ft.Colors.RED_700)
                for nome, qtd in abaixo_limite
            ],
            spacing=4
        ) if abaixo_limite else ft.Text("✔️ Estoque suficiente", size=12, color=ft.Colors.GREEN_700) 

        return ft.Container(
            col={"sm": 12, "md": 4},
            bgcolor="#FFFFFF",
            border_radius=12,
            shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
            padding=20,
            content=ft.Column([
                ft.Text("Estoque de Medicamentos", size=16, weight="bold", color="#111827"),
                ft.Text(f"Total em estoque: {total_estoque} unidades", size=14, weight="bold", color="#1E3A8A"),
                ft.Text("Medicamentos com unidades menores que 5:", size=13, color="#6B7280"),
                lista_criticos
            ], spacing=10)
        )

    def menu_item(self, icon, text, on_click=None):
        item = ft.Container(
            content=ft.Row([
                ft.Icon(icon, color="#059669", size=24),
                ft.Text(text, size=14, visible=self.sidebar_open, color="#374151")
            ], spacing=10, alignment=ft.MainAxisAlignment.START),
            padding=ft.padding.symmetric(vertical=12, horizontal=12),
            bgcolor="#FFFFFF",
            border_radius=10,
            ink=True,
            animate=ft.Animation(200, "easeInOut"),
            on_click=on_click
        )
        def on_hover(e):
            item.bgcolor = "#d1eefa" if e.data == "true" else "#FFFFFF"
            item.update()
        item.on_hover = on_hover
        return item

    def side_menu(self):
            def create_menu_item(icon, text, on_click=None):
                container = ft.Container(
                    padding=ft.padding.symmetric(vertical=12, horizontal=10),
                    content=ft.Row(
                        [
                            ft.Icon(icon, color=ft.Colors.BLUE_600, size=24),
                            ft.Text(text, size=16, visible=self.sidebar_open, color="#111827"),
                        ],
                        alignment=ft.MainAxisAlignment.START,
                        spacing=15,
                    ),
                    ink=True,
                    border_radius=8,
                    bgcolor="#FFFFFF",
                    on_click=on_click,
                    animate=ft.Animation(200, "easeInOut"),
                    margin=ft.margin.only(bottom=8),
                )

                def on_hover(e):
                    container.bgcolor = "#d1eefa" if e.data == "true" else "#FFFFFF"
                    container.update()

                container.on_hover = on_hover
                return container

            menu_items = [
                create_menu_item(ft.Icons.HOME_OUTLINED, "Início", self.load_dashboard),
                create_menu_item(ft.Icons.CALENDAR_MONTH_OUTLINED, "Agendamentos", self.load_agendamentos),
                create_menu_item(ft.Icons.MEDICAL_SERVICES_OUTLINED, "Medicamentos", self.load_medicamentos),
                create_menu_item(ft.Icons.PERSON_OUTLINED, "Pacientes", self.load_pacientes),
            ]

            botao_sair = ft.Container(
                padding=ft.padding.symmetric(vertical=12, horizontal=10),
                content=ft.Row(
                    [
                        ft.Icon(ft.Icons.LOGOUT, color="#DC2626", size=24),
                        ft.Text("Sair", size=16, visible=self.sidebar_open, color="#DC2626"),
                    ],
                    alignment=ft.MainAxisAlignment.START,
                    spacing=15,
                ),
                border_radius=8,
                bgcolor="#FEE2E2",
                ink=True,
                on_click=lambda e: self.page.go("/escolha_usuario"),
                animate=ft.Animation(200, "easeInOut"),
            )

            def on_hover_sair(e):
                botao_sair.bgcolor = "#FCA5A5" if e.data == "true" else "#FEE2E2"
                botao_sair.update()

            botao_sair.on_hover = on_hover_sair

            return ft.Container(
                width=240 if self.sidebar_open else 80,
                bgcolor="#F9FAFB",
                border=ft.border.only(right=ft.BorderSide(1, "#E5E7EB")),
                padding=ft.padding.symmetric(horizontal=12, vertical=20),
                animate=ft.Animation(duration=250, curve=ft.AnimationCurve.EASE_OUT),
                content=ft.Column(
                    expand=True,
                    spacing=10,
                    controls=[
                        ft.Container(
                            alignment=ft.alignment.center,
                            padding=ft.padding.only(bottom=10),
                            content=ft.Image(src="img/logo.png", width=80, height=80)
                        ),
                        ft.Divider(thickness=1),
                        *menu_items,
                        ft.Container(expand=True),  # O espaçador mágico
                        botao_sair
                    ]
                )
            )

    def header(self):
        return ft.Container(
            padding=ft.padding.symmetric(horizontal=20, vertical=14),
            bgcolor="white",
            border=ft.border.only(bottom=ft.BorderSide(1, "#E5E7EB")),
            content=ft.ResponsiveRow(
                columns=12,
                spacing=10,
                run_spacing=10,
                controls=[
                    ft.Container(
                        col={"sm": 12, "md": 6},
                        alignment=ft.alignment.center_left,
                        content=ft.Text(
                            "FarmConnect - Painel de Controle",
                            size=22,
                            weight="bold",
                            color=ft.Colors.BLUE_900,
                            text_align=ft.TextAlign.START
                        )
                    ),
                    ft.Container(
                        col={"sm": 12, "md": 6},
                        alignment=ft.alignment.center_right,
                        content=ft.Column([
                            ft.Row(
                                wrap=True,
                                spacing=10,
                                controls=[
                                    ft.Text("Bem-vindo!", size=20, color=ft.Colors.BLUE_900),
                                    ft.Text(self.page.session.get("admin_nome") or "Administrador", size=20, weight="bold", color=ft.Colors.BLUE_900),
                                    
                                ]
                            )
                        ])
                    )
                ]
            )
        )

    def graph_cards(self):
        return ft.ResponsiveRow(
            columns=12,
            spacing=20,
            controls=[
                ft.Container(
                    col={"sm": 12, "md": 4},
                    bgcolor="#FFFFFF",
                    border_radius=12,
                    shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                    padding=20,
                    content=ft.Column([
                        ft.Row([ft.Text("Agendamentos Pendentes", size=16, weight="bold", color="#111827")]),
                        ft.Text(str(self.count_agendamentos_pendentes()), size=40, weight="bold", color="#111827"),
                        ft.Text("Aguardando aprovação", size=13, color="#111827"),
                        ft.Container(height=12),
                        ft.OutlinedButton(
                            text="Ver Agendamentos",
                            icon=ft.Icons.CALENDAR_MONTH,
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                            on_click=self.load_agendamentos
                        )
                    ], spacing=10)
                ),

                self.card_estoque_medicamentos(),

                ft.Container(
                    col={"sm": 12, "md": 4},
                    bgcolor="#FFFFFF",
                    border_radius=12,
                    shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                    padding=20,
                    content=self.card_medicamentos_mais_solicitados()
                )
            ]
        )
    
    def card_medicamentos_mais_solicitados(self):
        dados = medicamentos_mais_solicitados()
        cores = ["#047857", "#059669", "#10B981", "#3B82F6", "#6366F1"]

        if not dados:
            return ft.Column([
                ft.Text("Medicamentos mais solicitados", size=16, weight="bold", color="#111827"),
                ft.Text("Nenhum agendamento registrado.", size=13, color="#6B7280")
            ])

        return ft.Column([
            ft.Text("Medicamentos mais solicitados", size=16, weight="bold", color="#111827"),
            ft.Column([
                ft.Row([
                    ft.Container(width=50 - i * 5, height=20, bgcolor=cores[i % len(cores)], border_radius=5),
                    ft.Text(f"{nome} ({quantidade})", size=13)
                ], spacing=8)
                for i, (nome, quantidade) in enumerate(dados)
            ], spacing=6)
        ])



    def metric_cards(self):
        return ft.ResponsiveRow(
            columns=12,
            spacing=20,
            controls=[
                ft.Container(
                    col={"sm": 12, "md": 4},
                    bgcolor="#FFFFFF",
                    border_radius=12,
                    shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                    padding=20,
                    content=ft.Column([
                        ft.Text("Pacientes Cadastrados", size=14, weight="bold", color="#111827"),
                        ft.Text(len(listar_usuarios()), size=34, weight="bold", color="#111827"),
                        ft.Text("Valor dinâmico", size=12, color="#6B7280"),
                        ft.OutlinedButton("+ Adicionar Paciente", style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)), on_click=self.load_cadastro_paciente)
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER)
                ),
                ft.Container(
                    col={"sm": 12, "md": 4},
                    bgcolor="#FFFFFF",
                    border_radius=12,
                    shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                    padding=20,
                    content=ft.Column([
                        ft.Text("Agendamentos Feitos Hoje", size=14, weight="bold"),
                        ft.Text(str(self.count_agendamentos_do_dia()), size=34, weight="bold", color="#111827"),
                        ft.Text("Valor dinâmico", size=12, color="#6B7280"),
                        ft.OutlinedButton(
                            "+ Novo Agendamento",
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                            on_click=self.load_cadastro_agendamento  # <- Aqui você adiciona a função
                        )
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER)
                ),
                ft.Container(
                    col={"sm": 12, "md": 4},
                    bgcolor="#FFFFFF",
                    border_radius=12,
                    shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26),
                    padding=20,
                    content=ft.Column([
                        ft.Text("Medicamentos Cadastrados", size=14, weight="bold", color="#111827"),
                        ft.Text(len(listar_medicamentos()), size=34, weight="bold", color="#111827"),
                        ft.Text("Valor dinâmico", size=12, color="#6B7280"),
                        ft.OutlinedButton(
                            "+ Adicionar Medicamento",
                            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=8)),
                            on_click=self.load_cadastro_medicamento  # <- Aqui você adiciona a função
                        )
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER)
                )
            ]
        )

    def load_dashboard(self, e=None):
        self.current_view.controls = [
            ft.Container(
                padding=20,
                content=ft.Column([
                    self.graph_cards(),
                    ft.Container(height=20),
                    self.metric_cards(),
                    ft.Container(height=40),

                    # Botões centralizados para baixar relatórios
                    ft.Container(
                        alignment=ft.alignment.center,
                        padding=20,
                        content=ft.Row(
                            alignment=ft.MainAxisAlignment.CENTER,
                            spacing=30,
                            controls=[
                                ft.ElevatedButton(
                                    text=" Baixar Relatório (PDF) ",
                                    icon=ft.Icons.PICTURE_AS_PDF,
                                    bgcolor=ft.Colors.BLUE_900,
                                    color=ft.Colors.WHITE,
                                    height=55,
                                    style=ft.ButtonStyle(
                                        shape=ft.RoundedRectangleBorder(radius=14),
                                        padding=ft.padding.symmetric(vertical=12),
                                        elevation=6
                                    ),
                                    on_click=lambda e: self.gerar_relatorio_pdf()
                                ),
                                ft.ElevatedButton(
                                    text=" Baixar Relatório (Excel) ",
                                    icon=ft.Icons.TABLE_VIEW,
                                    bgcolor=ft.Colors.GREEN_700,
                                    color=ft.Colors.WHITE,
                                    height=55,
                                    style=ft.ButtonStyle(
                                        shape=ft.RoundedRectangleBorder(radius=14),
                                        padding=ft.padding.symmetric(vertical=12),
                                        elevation=6
                                    ),
                                    on_click=lambda e: self.gerar_relatorio_excel()
                                )
                            ]
                        )
                    )
                ], spacing=20)
            )
        ]
        self.page.update()


    def gerar_rows_medicamentos(self, lista):
        rows = []
        for med in lista:
            inativo = med[11] == 0  # Índice 9 é o campo 'ativo' na tabela
            estilo_riscado = ft.TextStyle(
                decoration=ft.TextDecoration.LINE_THROUGH,
                color=ft.Colors.RED
            ) if inativo else None

            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(med[0]), style=estilo_riscado)),  # ID
                        ft.DataCell(ft.Text(med[1], style=estilo_riscado)),       # Nome
                        ft.DataCell(ft.Text(med[2] or "-", style=estilo_riscado)),# Código
                        ft.DataCell(ft.Text(med[6] or "-", style=estilo_riscado)),# Categoria
                        ft.DataCell(ft.Text(med[7] or "-", style=estilo_riscado)),# Fabricante
                        ft.DataCell(ft.Text(med[8] or "-", style=estilo_riscado)),# Farmácia
                        ft.DataCell(ft.Text(str(med[5] or 0), style=estilo_riscado)),  # Estoque
                        ft.DataCell(
                            ft.IconButton(
                                icon=ft.Icons.LOCK_OPEN if inativo else ft.Icons.CANCEL_OUTLINED,
                                icon_color=ft.Colors.GREEN_400 if inativo else ft.Colors.RED,
                                tooltip="Reativar" if inativo else "Desativar",
                                icon_size=22,
                                style=ft.ButtonStyle(padding=0),
                                on_click=(lambda e, id=med[0], inativo=med[11]:
                                    self.reativar_medicamento(id) if inativo == 0 else self.desativar_medicamento(id)
                                )
                            )
                        )
                    ]
                )
            )
        return rows


    def load_medicamentos(self, e=None, medicamento=None):

        self.current_view.controls.clear()
        self.editando_medicamento = medicamento is not None

        if not hasattr(self, "cancelados"):
            self.cancelados = set()

        categorias = listar_categorias()
        fabricantes = listar_fabricantes()

        if medicamento:
            self.medicamento_atual = medicamento

        medicamentos = listar_medicamentos()

        self.campo_busca_medicamento = ft.TextField(
                                        hint_text="Buscar medicamento...",
                                        prefix_icon=ft.Icons.SEARCH,
                                        border_radius=30,
                                        bgcolor="#F9FAFB",
                                        height=50,
                                        on_change=self.filtrar_medicamentos,
                                    )

        self.tabela_medicamentos = ft.DataTable(
            ref=self.tabela_medicamentos_ref,
            heading_row_color="#F9FAFB",
            border=ft.border.all(1, "#E5E7EB"),
            border_radius=12,
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("Código")),
                ft.DataColumn(ft.Text("Categoria")),
                ft.DataColumn(ft.Text("Fabricante")),
                ft.DataColumn(ft.Text("Farmácia")),
                ft.DataColumn(ft.Text("Estoque")),
                ft.DataColumn(ft.Text("Ações")),
            ],
            rows=self.gerar_rows_medicamentos(medicamentos)
        )

        # Painel lateral opcional (se estiver editando)
        self.campo_nome = ft.TextField(label="Nome do Medicamento", value=medicamento.get("nome") if medicamento else "")
        self.campo_codigo = ft.TextField(label="Código do Medicamento", value=medicamento.get("codigo") if medicamento else "")
        self.dropdown_fabricante = ft.Dropdown(
            label="Fabricante",
            border_radius=10,
            bgcolor="#F0FDF4",
            options=[ft.dropdown.Option(str(f[0]), f[1]) for f in fabricantes],
            value=str(medicamento.get("fabricante_id")) if medicamento else None,
            expand=True
        )
        self.dropdown_categoria = ft.Dropdown(
            label="Categoria",
            border_radius=10,
            bgcolor="#F0FDF4",
            options=[ft.dropdown.Option(str(c[0]), c[1]) for c in categorias],
            value=str(medicamento.get("categoria_id")) if medicamento else None,
            expand=True
        )
        self.campo_descricao = ft.TextField(label="Descrição", value=medicamento.get("descricao") if medicamento else "")
        self.campo_imagem = ft.TextField(label="URL da Imagem", value=medicamento.get("imagem") if medicamento else "", border_radius=10)
        self.campo_estoque = ft.TextField(label="Estoque Atual", keyboard_type=ft.KeyboardType.NUMBER, value=str(medicamento.get("estoque")) if medicamento else "")

        detalhes_medicamento = ft.Container(
            bgcolor="white",
            border_radius=12,
            padding=20,
            shadow=ft.BoxShadow(blur_radius=8, color="#E2E8F0"),
            expand=1,
            visible=self.editando_medicamento,
            animate_opacity=300,
            opacity=1.0 if self.editando_medicamento else 0.0,
            content=ft.Column([
                ft.Text("Detalhes do Medicamento", size=20, weight="bold", color="#059669"),
                ft.Divider(),
                self.campo_nome,
                self.campo_codigo,
                self.dropdown_fabricante,
                self.dropdown_categoria,
                self.campo_descricao,
                self.campo_imagem,
                self.campo_estoque,
                ft.TextField(label="Observações", multiline=True, min_lines=3, max_lines=5),
                ft.Container(height=20),
                ft.Row([
                    ft.ElevatedButton(
                        "Salvar",
                        bgcolor="#059669",
                        color="white",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=self.salvar_medicamento
                    ),
                    ft.OutlinedButton(
                        "Cancelar",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=lambda e: self.load_medicamentos(),
                    ),
                ], spacing=10),
            ], spacing=12)
        )

        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column([
                    ft.Row([
                        ft.Icon(name=ft.Icons.LOCAL_PHARMACY, size=40, color=ft.Colors.BLUE_600),
                        ft.Text("Gerenciamento de Medicamentos", size=32, weight="bold", color=ft.Colors.BLUE_900),
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER),
                    ft.Text(
                        "Adicione, edite ou remova medicamentos disponíveis no sistema.",
                        size=14,
                        color=ft.Colors.GREY_700,
                        text_align=ft.TextAlign.CENTER
                    ),
                    ft.Divider(height=30),

                    ft.Container(
                        bgcolor="#FFFFFF",
                        border_radius=20,
                        padding=25,
                        shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
                        content=ft.Column([
                            ft.Row([
                                ft.Text("📋 Lista de Medicamentos", size=20, weight="bold", color="#111827"),
                                ft.IconButton(
                                    icon=ft.Icons.ADD,
                                    tooltip="Adicionar novo medicamento",
                                    icon_color=ft.Colors.BLUE_600,
                                    on_click=self.load_cadastro_medicamento
                                ),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Divider(),
                            ft.Row([
                                ft.Container(
                                    expand=True,
                                    content=self.campo_busca_medicamento
                                    )
                            ]),
                            ft.Container(height=20),

                            ft.ResponsiveRow(
                                
                                columns=12,
                                spacing=20,
                                run_spacing=20,
                                controls=[
                                    ft.Container(
                                        col={"sm": 12, "md": 12},
                                        expand=True,
                                        content=ft.Row(
                                            scroll=ft.ScrollMode.ALWAYS,
                                            controls = [
                                                self.tabela_medicamentos,
                                            ],
                                        )
                                    ),
                                    ft.Container(
                                        col={"sm": 12, "md": 4},
                                        content=detalhes_medicamento
                                    )
                                ]
                            )
                        ], width = 1300, spacing=20)
                    )
                ], spacing=20)
            )
        )

        self.page.update()

    def filtrar_medicamentos(self, e):
        termo = self.campo_busca_medicamento.value.strip().lower()
        medicamentos = listar_medicamentos()
        resultado = [
            m for m in medicamentos
            if termo in (m[1] or "").lower() or termo in (m[2] or "").lower()
        ]

        self.tabela_medicamentos_ref.current.rows = self.gerar_rows_medicamentos(resultado)
        self.tabela_medicamentos_ref.current.update()

    def salvar_medicamento(self, e=None):
        nome = self.campo_nome.value.strip()
        codigo = self.campo_codigo.value.strip()
        descricao = self.campo_descricao.value.strip()
        imagem = self.imagem_escolhida
        estoque_str = self.campo_estoque.value.strip()
        
        categoria_id = int(self.dropdown_categoria.value) if self.dropdown_categoria.value else None
        fabricante_id = int(self.dropdown_fabricante.value) if self.dropdown_fabricante.value else None

        # Validação
        if not nome:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("O nome do medicamento é obrigatório."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        try:
            estoque = int(estoque_str)
        except ValueError:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Estoque deve ser um número válido."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        # Edição ou adição
        if self.editando_medicamento and hasattr(self, "medicamento_atual"):
            editar_medicamento(
                self.medicamento_atual["id"],
                nome=nome,
                codigo=codigo,
                descricao=descricao,
                imagem=imagem,
                estoque=estoque,
                categoria_id=categoria_id,
                fabricante_id=fabricante_id
            )
        else:
            farmacia_id = int(self.dropdown_farmacia.value) if self.dropdown_farmacia.value else None

            if not farmacia_id:
                self.page.snack_bar = ft.SnackBar(content=ft.Text("Você deve selecionar uma farmácia."), bgcolor="red")
                self.page.snack_bar.open = True
                self.page.update()
                return

            # 1. Cadastra o medicamento e pega o ID
            id_medicamento = adicionar_medicamento(
                nome=nome,
                codigo=codigo,
                descricao=descricao,
                imagem=imagem,
                estoque=estoque,
                categoria_id=categoria_id,
                fabricante_id=fabricante_id,
            )

            # 2. Relaciona ao estoque da farmácia
            adicionar_estoque(farmacia_id, id_medicamento, estoque)

        # Voltar para lista
        self.load_medicamentos()


    def load_cadastro_medicamento(self, e=None):
        self.editando_medicamento = False
        self.current_view.controls.clear()

        # Carregar categorias, fabricantes e farmácias
        categorias = listar_categorias()
        fabricantes = listar_fabricantes()
        farmacias = listar_farmacias()

        self.dropdown_categoria = ft.Dropdown(
            label="Categoria",
            border_radius=10,
            bgcolor="#F9FAFB",
            options=[ft.dropdown.Option(str(c[0]), c[1]) for c in categorias],
            expand=True
        )

        self.dropdown_fabricante = ft.Dropdown(
            label="Fabricante",
            border_radius=10,
            bgcolor="#F9FAFB",
            options=[ft.dropdown.Option(str(f[0]), f[1]) for f in fabricantes],
            expand=True
        )

        self.dropdown_farmacia = ft.Dropdown(
            label="Farmácia",
            options=[ft.dropdown.Option(str(f[0]), f"{f[1]} - {f[3]}") for f in farmacias],
            border_radius=10,
            bgcolor="#F9FAFB",
            expand=True
        )

        self.campo_nome = ft.TextField(label="Nome do Medicamento", border_radius=10, bgcolor="#F9FAFB")
        self.campo_codigo = ft.TextField(label="Código do Medicamento", border_radius=10, bgcolor="#F9FAFB")
        self.campo_estoque = ft.TextField(label="Quantidade em Estoque", keyboard_type=ft.KeyboardType.NUMBER, border_radius=10, bgcolor="#F9FAFB")
        self.campo_descricao = ft.TextField(label="Descrição do Medicamento", border_radius=10, bgcolor="#F9FAFB")
        self.imagem_selecionada = ft.Ref[ft.Image]()
        self.imagem_escolhida = "img/seringa.png"  # valor inicial

        def selecionar_imagem(imagem_path):
            def handler(e):
                self.imagem_escolhida = imagem_path
                self.imagem_selecionada.current.src = imagem_path
                self.imagem_selecionada.current.update()
            return handler

        self.container_selecao_imagem = ft.Container(
            content=ft.Column([
                ft.Text("Escolha a imagem do medicamento:", size=14),
                ft.Row([
                    ft.GestureDetector(
                        content=ft.Image(src="img/seringa.png", width=100, height=100),
                        on_tap=selecionar_imagem("img/seringa.png")
                    ),
                    ft.GestureDetector(
                        content=ft.Image(src="img/comprimido.png", width=100, height=100),
                        on_tap=selecionar_imagem("img/comprimido.png")
                    ),
                ], spacing=20),
                ft.Text("Imagem selecionada:"),
                ft.Image(ref=self.imagem_selecionada, src=self.imagem_escolhida, width=80, height=80)
            ], spacing=10)
        )

        self.campo_observacoes = ft.TextField(label="Observações", multiline=True, min_lines=3, max_lines=5, border_radius=10, bgcolor="#F9FAFB")

        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column(
                    [
                        ft.Text("Cadastrar Novo Medicamento", size=28, weight="bold", color=ft.Colors.BLUE_900),
                        ft.Container(height=20),
                        ft.Container(
                            padding=30,
                            bgcolor="#FFFFFF",
                            border_radius=20,
                            shadow=ft.BoxShadow(blur_radius=20, spread_radius=2, color="#CBD5E1", offset=ft.Offset(0, 8)),
                            content=ft.Column(
                                [
                                    self.campo_nome,
                                    self.campo_codigo,
                                    self.campo_descricao,
                                    ft.Row([
                                        self.dropdown_categoria,
                                        ft.IconButton(
                                            icon=ft.Icons.ADD_CIRCLE_OUTLINE,
                                            tooltip="Cadastrar nova categoria",
                                            icon_color=ft.Colors.BLUE_900,
                                            on_click=self.load_cadastro_categoria
                                        )
                                    ], spacing=10),
                                    ft.Row([
                                        self.dropdown_fabricante,
                                        ft.IconButton(
                                            icon=ft.Icons.ADD_CIRCLE_OUTLINE,
                                            tooltip="Cadastrar novo fabricante",
                                            icon_color=ft.Colors.BLUE_900,
                                            on_click=self.load_cadastro_fabricante
                                        )
                                    ], spacing=10),
                                    self.dropdown_farmacia,
                                    self.campo_estoque,
                                    self.container_selecao_imagem,
                                    self.campo_observacoes,
                                    ft.Container(height=20),
                                    ft.Row(
                                        [
                                            ft.ElevatedButton(
                                                "Salvar Medicamento",
                                                bgcolor=ft.Colors.BLUE_600,
                                                color="white",
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                expand=True,
                                                on_click=self.salvar_medicamento
                                            ),
                                            ft.OutlinedButton(
                                                "Cancelar",
                                                expand=True,
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                on_click=lambda e: self.load_medicamentos(),
                                            ),
                                        ],
                                        spacing=20,
                                    )
                                ],
                                spacing=15,
                            ),
                        )
                    ],
                    spacing=20,
                    expand=True,
                    alignment=ft.MainAxisAlignment.CENTER,
                ),
            )
        )

        self.page.update()


    def load_cadastro_categoria(self, e=None):
        from database import adicionar_categoria

        campo_nome_categoria = ft.TextField(label="Nome da Categoria", border_radius=10, bgcolor="#F9FAFB")

        def salvar(e):
            nome = campo_nome_categoria.value.strip()
            if nome:
                adicionar_categoria(nome)
                self.load_cadastro_medicamento()
            else:
                self.page.snack_bar = ft.SnackBar(content=ft.Text("Nome da categoria é obrigatório"), bgcolor="red")
                self.page.snack_bar.open = True
                self.page.update()

        self.current_view.controls.clear()
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column([
                    ft.Text("Cadastrar Nova Categoria", size=26, weight="bold", color=ft.Colors.BLUE_900),
                    ft.Container(height=20),
                    campo_nome_categoria,
                    ft.Container(height=20),
                    ft.Row([
                        ft.ElevatedButton("Salvar", bgcolor=ft.Colors.BLUE_600, color="white", on_click=salvar, expand=True),
                        ft.OutlinedButton("Cancelar", on_click=lambda e: self.load_cadastro_medicamento(), expand=True),
                    ], spacing=20)
                ], spacing=10)
            )
        )

        self.page.update()

    
    def load_cadastro_fabricante(self, e=None):
        from database import adicionar_fabricante

        campo_nome_fabricante = ft.TextField(label="Nome do Fabricante", border_radius=10, bgcolor="#F9FAFB")

        def salvar(e):
            nome = campo_nome_fabricante.value.strip()
            if nome:
                adicionar_fabricante(nome)
                self.load_cadastro_medicamento()
            else:
                self.page.snack_bar = ft.SnackBar(content=ft.Text("Nome do fabricante é obrigatório."), bgcolor="red")
                self.page.snack_bar.open = True
                self.page.update()

        self.current_view.controls.clear()
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column([
                    ft.Text("Cadastrar Novo Fabricante", size=26, weight="bold", color=ft.Colors.BLUE_900),
                    ft.Container(height=20),
                    campo_nome_fabricante,
                    ft.Container(height=20),
                    ft.Row([
                        ft.ElevatedButton("Salvar", bgcolor=ft.Colors.BLUE_600, color="white", on_click=salvar, expand=True),
                        ft.OutlinedButton("Cancelar", on_click=lambda e: self.load_cadastro_medicamento(), expand=True),
                    ], spacing=20)
                ], spacing=10)
            )
        )

        self.page.update()

    
    def load_agendamentos(self, e=None, agendamento=None):
        self.current_view.controls.clear()
        self.editando_agendamento = agendamento is not None
        self.agendamento_atual = agendamento if agendamento else None

        # Buscar farmácias do banco de dados
        agendamentos_db = listar_agendamentos()
        
        # Campos de edição
        self.campo_busca_agendamentos = ft.TextField(
            hint_text="Buscar agendamentos...",
            prefix_icon=ft.Icons.SEARCH,
            border_radius=30,
            bgcolor="#F9FAFB",
            height=50,
            on_change=self.filtrar_agendamentos
        )
        
        self.current_view.controls.clear()

        # Campos do formulário
        self.campo_paciente = ft.TextField(label="Nome do Paciente", border_radius=10, bgcolor="#F9FAFB")
        self.campo_medicamento = ft.TextField(label="Nome do Medicamento", border_radius=10, bgcolor="#F9FAFB")
        self.campo_codigo = ft.TextField(label="Código do Medicamento", border_radius=10, bgcolor="#F9FAFB")
        self.campo_quantidade = ft.TextField(label="Quantidade", border_radius=10, bgcolor="#F9FAFB", keyboard_type=ft.KeyboardType.NUMBER)
        self.campo_data = ft.TextField(label="Data (AAAA-MM-DD)", border_radius=10, bgcolor="#F9FAFB")
        self.campo_horario = ft.TextField(label="Horário (HH:MM)", border_radius=10, bgcolor="#F9FAFB")

        # Painel lateral de edição
        self.painel_detalhes_agendamento = ft.Container(
            bgcolor="#FFFFFF",
            border_radius=20,
            padding=25,
            shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
            visible=self.editando_agendamento,
            animate_opacity=300,
            opacity=1.0 if self.editando_agendamento else 0.0,
            expand=1,
            content=ft.Column([
                ft.Text("🏥 Editar Agendamento", size=20, weight="bold", color=ft.Colors.BLUE_900),
                ft.Divider(),
                self.campo_paciente,
                self.campo_medicamento,
                self.campo_codigo,
                self.campo_quantidade,
                self.campo_data,
                self.campo_horario,
                ft.Container(height=20),
                ft.Row([
                    ft.ElevatedButton(
                        "Salvar",
                        bgcolor=ft.Colors.BLUE_600,
                        color="white",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=self.salvar_agendamento  # Novo método para salvar
                    ),
                    ft.OutlinedButton(
                        "Cancelar",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=lambda e: self.load_agendamentos()
                    )
                ], spacing=10)
            ], spacing=12)
        )

        # Renderiza a tabela com farmácias do banco
        self.renderizar_tabela_agendamentos(agendamentos_db)

    def gerar_rows_agendamentos(self, lista):
        status_cores = {
            "Pendente": ("#D97706", "#FEF3C7"),
            "Confirmado": ("#15803D", "#D1FAE5"),
            "Cancelado": ("#DC2626", "#FEE2E2"),
            "Concluído/Retirado": ("#047857", "#D1FAE5"),
        }

        def status_badge(status):
            cor_texto, cor_fundo = status_cores.get(status, ("#6B7280", "#E5E7EB"))
            return ft.Row([
                ft.Container(width=8, height=8, bgcolor=cor_texto, border_radius=20),
                ft.Text(status, color=cor_texto, weight="bold")
            ], spacing=6, alignment=ft.MainAxisAlignment.CENTER)

        rows = []
        for a in lista:
            acoes = []

            # ✅ Confirmar agendamento
            acoes.append(ft.IconButton(
                icon=ft.Icons.CHECK_CIRCLE_OUTLINE,
                icon_color="#059669",
                tooltip="Confirmar Agendamento",
                icon_size=20,
                on_click=lambda e, ag_id=a[0]: self.confirmar_agendamento(ag_id),
                style=ft.ButtonStyle(padding=0),
            ))

            # 🔁 Confirmar retirada (se status permitir)
            if a[7] in ["Confirmado", "Pendente"]:
                acoes.append(ft.IconButton(
                    icon=ft.Icons.EXIT_TO_APP,
                    icon_color=ft.Colors.BLUE_900,
                    tooltip="Confirmar Retirada",
                    icon_size=20,
                    on_click=lambda e, ag_id=a[0]: self.confirmar_retirada(ag_id),
                    style=ft.ButtonStyle(padding=0),
                ))

            # ❌ Cancelar agendamento
            acoes.append(ft.IconButton(
                icon=ft.Icons.CANCEL_OUTLINED,
                icon_color="#DC2626",
                tooltip="Cancelar Agendamento",
                icon_size=20,
                on_click=lambda e, ag_id=a[0]: self.cancelar_agendamento(ag_id),
                style=ft.ButtonStyle(padding=0),
            ))

            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(a[0]))),
                        ft.DataCell(ft.Text(a[1])),
                        ft.DataCell(ft.Text(a[2])),
                        ft.DataCell(ft.Text(a[3])),
                        ft.DataCell(ft.Text(a[4])),
                        ft.DataCell(ft.Text(a[5])),
                        ft.DataCell(ft.Text(a[6])),
                        ft.DataCell(
                            ft.Container(
                                content=status_badge(a[7]),
                                padding=ft.padding.symmetric(horizontal=8, vertical=4),
                                bgcolor=status_cores.get(a[7], "#E5E7EB")[1],
                                border_radius=20,
                            )
                        ),
                        ft.DataCell(
                            ft.Container(
                                width=140,  # AUMENTA ESPAÇO PARA OS 3 ÍCONES
                                alignment=ft.alignment.center,
                                content=ft.Row(
                                    controls=acoes,
                                    spacing=6,
                                    alignment=ft.MainAxisAlignment.CENTER
                                )
                            )
                        )
                    ]
                )
            )

        return rows


    def renderizar_tabela_agendamentos(self, lista):
        self.tabela_agendamentos = ft.DataTable(
            ref=self.tabela_agendamentos_ref,
            heading_row_color="#F9FAFB",
            border=ft.border.all(1, "#E5E7EB"),
            border_radius=12,
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("Paciente")),
                ft.DataColumn(ft.Text("Medicamento")),
                ft.DataColumn(ft.Text("Farmácia")),
                ft.DataColumn(ft.Text("Código")),
                ft.DataColumn(ft.Text("Data")),
                ft.DataColumn(ft.Text("Horário")),
                ft.DataColumn(ft.Text("Status")),
                ft.DataColumn(ft.Text("Ações")),
            ],
            rows=self.gerar_rows_agendamentos(lista)
        )

        self.current_view.controls.clear()
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column([
                    ft.Row([
                        ft.Icon(name=ft.Icons.LOCAL_HOSPITAL, size=40, color=ft.Colors.BLUE_600),
                        ft.Text("Gerenciamento de Agendamentos", size=32, weight="bold", color=ft.Colors.BLUE_900)
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER),
                    ft.Text(
                        "Adicione, edite ou gerencie os agendamentos cadastrados no sistema.",
                        size=14,
                        color=ft.Colors.GREY_700,
                        text_align=ft.TextAlign.CENTER
                    ),
                    ft.Divider(height=30),

                    ft.Container(
                        bgcolor="#FFFFFF",
                        border_radius=20,
                        padding=25,
                        shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
                        content=ft.Column([
                            ft.Row([
                                ft.Text("📋 Lista de Agendamentos", size=20, weight="bold", color="#111827"),
                                ft.IconButton(
                                    icon=ft.Icons.ADD,
                                    tooltip="Adicionar novo agendamento",
                                    icon_color=ft.Colors.BLUE_600,
                                    on_click=self.load_cadastro_agendamento
                                ),
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Divider(),
                            ft.Row([
                                ft.Container(
                                    expand=True,
                                    content=self.campo_busca_agendamentos
                                )
                            ]),
                            ft.Container(height=20),

                            ft.Row(
                                scroll=ft.ScrollMode.ALWAYS,
                                controls=[
                                        self.tabela_agendamentos,
                                ]
                            )
                        ], width=1300, spacing=20)
                    )
                ], spacing=20)
            )
        )

        self.page.update()


    def cnpj_change(self, e):
        texto_original = self.campo_cnpj.value
        numeros = ''.join(filter(str.isdigit, texto_original))[:14]

        if len(numeros) == 14:
            self.campo_cidade.focus()

    def cnpj_blur(self, e):
        texto_original = self.campo_cnpj.value
        numeros = ''.join(filter(str.isdigit, texto_original))[:14]
        formatado = ""

        if len(numeros) >= 2:
            formatado += numeros[:2] + "."
        if len(numeros) >= 5:
            formatado += numeros[2:5] + "."
        if len(numeros) >= 8:
            formatado += numeros[5:8] + "/"
        if len(numeros) >= 12:
            formatado += numeros[8:12] + "-"
        if len(numeros) > 12:
            formatado += numeros[12:]

        self.campo_cnpj.value = formatado
        self.campo_cnpj.update()
    
    def telefone_change(self, e):
        texto_original = self.campo_telefone.value
        numeros = ''.join(filter(str.isdigit, texto_original))[:13]

        if len(numeros) == 13:
            self.campo_telefone.focus()

    def telefone_blur(self, e):
        texto_original = self.campo_telefone.value
        numeros = ''.join(filter(str.isdigit, texto_original))[:11]
        formatado = ""

        if len(numeros) >= 1:
            formatado += "(" + numeros[:2] + ") "
        if len(numeros) >= 7:
            formatado += numeros[2:7] + "-"
        if len(numeros) > 7:
            formatado += numeros[7:]
        elif len(numeros) > 2:
            formatado += numeros[2:7]

        self.campo_telefone.value = formatado
        self.campo_telefone.update()

    def load_pacientes(self, e=None, paciente=None):
        self.current_view.controls.clear()
        self.editando_paciente = paciente is not None
        self.paciente_atual = paciente if paciente else None

        # Buscar pacientes do banco de dados
        pacientes_db = listar_usuarios()

        def telefone_change_paciente(e):
            numeros = ''.join(filter(str.isdigit, self.campo_telefone_paciente.value))[:11]
            if len(numeros) == 11:
                self.campo_senha.focus()

        def telefone_blur_paciente(e):
            numeros = ''.join(filter(str.isdigit, self.campo_telefone_paciente.value))[:11]
            fmt = ""
            if len(numeros) >= 1:
                fmt += "(" + numeros[:2] + ") "
            if len(numeros) >= 7:
                fmt += numeros[2:7] + "-"
            if len(numeros) > 7:
                fmt += numeros[7:]
            elif len(numeros) > 2:
                fmt += numeros[2:7]
            self.campo_telefone_paciente.value = fmt
            self.campo_telefone_paciente.update()

        self.campo_telefone_paciente = ft.TextField(
            label="Telefone",
            on_blur=telefone_blur_paciente,
            on_change=telefone_change_paciente,
            border_radius=10,
            bgcolor="#F9FAFB"
        )

        def cpf_blur(e):
            texto_original = self.campo_cpf.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:11]

            formatado = ""
            if len(numeros) >= 3:
                formatado += numeros[:3] + "."
            if len(numeros) >= 6:
                formatado += numeros[3:6] + "."
            if len(numeros) >= 9:
                formatado += numeros[6:9] + "-"
            if len(numeros) > 9:
                formatado += numeros[9:]
            elif len(numeros) > 6:
                formatado += numeros[6:9]
            elif len(numeros) > 3:
                formatado += numeros[3:6]
            elif len(numeros) > 0:
                formatado += numeros[0:3]

            self.campo_cpf.value = formatado
            self.campo_cpf.update()

        def cpf_change(e):
            texto_original = self.campo_cpf.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:11]

            if len(numeros) == 11:
                self.campo_nascimento.focus()

        def nascimento_change(e):
            texto_original = self.campo_nascimento.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:8]

            if len(numeros) == 8:
                pass

        def nascimento_blur(e):
            texto_original = self.campo_nascimento.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:8]
            formatado = ""
            if len(numeros) >= 2:
                formatado += numeros[:2] + "/"
            if len(numeros) >= 4:
                formatado += numeros[2:4] + "/"
            if len(numeros) > 4:
                formatado += numeros[4:]
            elif len(numeros) > 2:
                formatado += numeros[2:]
            self.campo_nascimento.value = formatado
            self.campo_nascimento.update()

        self.current_view.controls.clear()
        
        # Campos de edição
        self.campo_busca_pacientes = ft.TextField(
            hint_text="Buscar paciente...",
            prefix_icon=ft.Icons.SEARCH,
            border_radius=30,
            bgcolor="#F9FAFB",
            height=50,
            on_change=self.filtrar_pacientes
        )

        self.campo_nome_paciente = ft.TextField(label="Nome do Paciente", border_radius=10, bgcolor="#F9FAFB")
        self.campo_cpf = ft.TextField(label="CPF", on_blur=cpf_blur, on_change=cpf_change, border_radius=10, bgcolor="#F9FAFB")
        self.campo_email = ft.TextField(label="Email", border_radius=10, bgcolor="#F9FAFB")
        self.campo_nascimento = ft.TextField(label="Data de Nascimento (DD/MM/AAAA)", on_blur=nascimento_blur, on_change=nascimento_change, border_radius=10, bgcolor="#F9FAFB")
        self.campo_senha = ft.TextField(label="Senha", password=True, can_reveal_password=True, border_radius=10, bgcolor="#F9FAFB")
        self.campo_confirmar_senha = ft.TextField(label="Confirmar Senha", password=True, can_reveal_password=True, border_radius=10, bgcolor="#F9FAFB")

        # Painel lateral de edição
        self.painel_detalhes_paciente = ft.Container(
            bgcolor="#FFFFFF",
            border_radius=20,
            padding=25,
            shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
            visible=self.editando_paciente,
            animate_opacity=300,
            opacity=1.0 if self.editando_paciente else 0.0,
            expand=1,
            content=ft.Column([
                ft.Text("🏥 Adicionar Paciente", size=20, weight="bold", color=ft.Colors.BLUE_900),
                ft.Divider(),
                self.campo_nome_paciente,
                self.campo_email,
                self.campo_cpf,
                self.campo_nascimento,
                self.campo_telefone_paciente,
                self.campo_senha,
                self.campo_confirmar_senha,
                ft.Container(height=20),
                ft.Row([
                    ft.ElevatedButton(
                        "Salvar",
                        bgcolor=ft.Colors.BLUE_600,
                        color="white",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=self.salvar_paciente  # Novo método para salvar
                    ),
                    ft.OutlinedButton(
                        "Cancelar",
                        expand=True,
                        style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                        on_click=lambda e: self.load_pacientes()
                    )
                ], spacing=10)
            ], spacing=12)
        )

        # Renderiza a tabela com farmácias do banco
        self.renderizar_tabela_pacientes(pacientes_db)

    def gerar_rows_pacientes(self, lista):
        from datetime import datetime

        def fmt_data_nasc(s: str) -> str:
            if not s:
                return "-"
            s = s.strip()

            # 1) Tenta formatos explícitos primeiro
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d"):
                try:
                    return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
                except:
                    pass

            # 2) Fallback por dígitos
            dig = "".join(ch for ch in s if ch.isdigit())
            if len(digitos := dig) == 8:
                # Se parece começar com ano (19xx/20xx) ou a string original começa com ano
                if s.startswith(("19", "20")) or (digitos[:2] in {"19", "20"}):
                    yyyy, mm, dd = digitos[0:4], digitos[4:6], digitos[6:8]
                else:
                    dd, mm, yyyy = digitos[0:2], digitos[2:4], digitos[4:8]
                return f"{dd}/{mm}/{yyyy}"
            return s  # mantém como veio se nada casar

        def fmt_data_criacao(s: str) -> str:
            if not s:
                return "-"
            s = s.strip()
            for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"):
                try:
                    return datetime.strptime(s, fmt).strftime("%d/%m/%Y %H:%M:%S")
                except:
                    pass
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y"):
                try:
                    return datetime.strptime(s, fmt).strftime("%d/%m/%Y")
                except:
                    pass
            return s

        status_cores = {
            "Pendente": ("#D97706", "#FEF3C7"),
            "Aprovado": ("#15803D", "#D1FAE5"),
            "Cancelado": ("#DC2626", "#FEE2E2"),
            "Recusado": ("#DC2626", "#FEE2E2"),
        }

        def status_badge(status):
            cor_texto, cor_fundo = status_cores.get(status, ("#6B7280", "#E5E7EB"))
            return ft.Row(
                [ft.Container(width=8, height=8, bgcolor=cor_texto, border_radius=20),
                ft.Text(status, color=cor_texto, weight="bold")],
                spacing=6, alignment=ft.MainAxisAlignment.CENTER
            )

        rows = []
        for p in lista:
            rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(p[0]))),                 # ID
                        ft.DataCell(ft.Text(p[1])),                      # Nome
                        ft.DataCell(ft.Text(p[3])),                      # CPF
                        ft.DataCell(ft.Text(fmt_data_nasc(p[4]))),       # Nascimento -> DD/MM/AAAA
                        ft.DataCell(ft.Text(fmt_data_criacao(p[6]))),    # Criação -> DD/MM/AAAA HH:MM:SS
                        ft.DataCell(
                            ft.Container(
                                content=status_badge(p[7]),
                                padding=ft.padding.symmetric(horizontal=4, vertical=4),
                                bgcolor=status_cores.get(p[7], "#E5E7EB")[1],
                                border_radius=20,
                            )
                        ),
                        ft.DataCell(
                            ft.Row(
                                [
                                    ft.IconButton(
                                        icon=ft.Icons.CHECK_CIRCLE_OUTLINE,
                                        icon_color="#059669",
                                        tooltip="Aprovar paciente",
                                        on_click=lambda e, pid=p[0]: self.aprovar_usuario(pid),
                                    ),
                                    ft.IconButton(
                                        icon=ft.Icons.CANCEL_OUTLINED,
                                        icon_color="#DC2626",
                                        tooltip="Recusar paciente",
                                        on_click=lambda e, pid=p[0]: self.recusar_usuario(pid),
                                    ),
                                ],
                                spacing=8,
                            )
                        ),
                    ]
                )
            )
        return rows

    def renderizar_tabela_pacientes(self, lista):
        self.tabela_pacientes = ft.DataTable(
            ref=self.tabela_pacientes_ref,
            heading_row_color="#F9FAFB",
            border=ft.border.all(1, "#E5E7EB"),
            border_radius=12,
            columns=[
                ft.DataColumn(ft.Text("ID")),
                ft.DataColumn(ft.Text("Nome")),
                ft.DataColumn(ft.Text("CPF")),
                ft.DataColumn(ft.Text("Nascimento")),
                ft.DataColumn(ft.Text("Data de Criação")),
                ft.DataColumn(ft.Text("Status")),
                ft.DataColumn(ft.Text("Ações")),
            ],
            rows= self.gerar_rows_pacientes(lista)
        )
        #Teste
        self.current_view.controls.clear()
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column([
                    ft.Row([
                        ft.Icon(name=ft.Icons.LOCAL_HOSPITAL, size=40, color=ft.Colors.BLUE_600),
                        ft.Text("📋 Gerenciamento de Pacientes", size=32, weight="bold", color=ft.Colors.BLUE_900)
                    ], spacing=10, alignment=ft.MainAxisAlignment.CENTER),
                    ft.Text("Adicione, edite ou gerencie os pacientes cadastrados no sistema.",
                            size=14, color=ft.Colors.GREY_700, text_align=ft.TextAlign.CENTER),
                    ft.Divider(height=30),

                    ft.Container(
                        bgcolor="#FFFFFF",
                        border_radius=20,
                        padding=25,
                        expand=True,
                        shadow=ft.BoxShadow(blur_radius=20, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
                        content=ft.Column([
                            ft.Row([
                                ft.Text("📋 Lista de Pacientes", size=20, weight="bold", color="#111827"),
                                ft.IconButton(
                                    icon=ft.Icons.ADD,
                                    tooltip="Adicionar novo paciente",
                                    icon_color=ft.Colors.BLUE_600,
                                    on_click=lambda e: self.load_pacientes(paciente={})
                                )
                            ], alignment=ft.MainAxisAlignment.SPACE_BETWEEN),
                            ft.Divider(),

                            ft.Row([
                                ft.Container(
                                    expand=True,
                                    content=self.campo_busca_pacientes
                                )
                            ]),
                            ft.Container(height=20),

                            ft.Container(
                                expand=True,
                                content=ft.Stack(
                                    expand=True,
                                    controls=[
                                        ft.Row(
                                            scroll=ft.ScrollMode.ALWAYS,
                                            controls=[
                                                self.tabela_pacientes
                                            ]
                                        ),
                                        ft.AnimatedSwitcher(
                                            content=self.painel_detalhes_paciente if self.editando_paciente else ft.Container(),
                                            transition=ft.Animation(300, "easeInOut")
                                        )
                                    ]
                                )
                            )
                        ], width=1100, spacing=20)
                    )
                ], spacing=20)
            )
        )

        self.page.update()

    def filtrar_pacientes(self, e):
        termo = self.campo_busca_pacientes.value.strip().lower()
        pacientes = listar_usuarios()
        resultado = [p for p in pacientes if termo in p[1].lower() or termo in p[3].lower()]

        self.tabela_pacientes_ref.current.rows = self.gerar_rows_pacientes(resultado)
        self.tabela_pacientes_ref.current.update()

    # Funções para Aprovar e Recusar Pacientes
    def aprovar_usuario(self, paciente_id):
        aprovar_usuario(paciente_id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Paciente aprovado com sucesso!"), bgcolor="green")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_pacientes()

    def recusar_usuario(self, paciente_id):
        recusar_usuario(paciente_id)
        self.page.snack_bar = ft.SnackBar(content=ft.Text("Paciente recusado com sucesso!"), bgcolor="red")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_pacientes()

    def load_cadastro_paciente(self, e=None):
        def cpf_blur(e):
            texto_original = self.campo_cpf.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:11]

            formatado = ""
            if len(numeros) >= 3:
                formatado += numeros[:3] + "."
            if len(numeros) >= 6:
                formatado += numeros[3:6] + "."
            if len(numeros) >= 9:
                formatado += numeros[6:9] + "-"
            if len(numeros) > 9:
                formatado += numeros[9:]
            elif len(numeros) > 6:
                formatado += numeros[6:9]
            elif len(numeros) > 3:
                formatado += numeros[3:6]
            elif len(numeros) > 0:
                formatado += numeros[0:3]

            self.campo_cpf.value = formatado
            self.campo_cpf.update()

        def cpf_change_cadastro(e):
            texto_original = self.campo_cpf.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:11]

            if len(numeros) == 11:
                self.campo_nascimento.focus()
        
        def telefone_change_cadastro(e):
            numeros = ''.join(filter(str.isdigit, self.campo_telefone_paciente.value))[:11]
            if len(numeros) == 11:
                self.campo_senha.focus()

        def telefone_blur_cadastro(e):
            numeros = ''.join(filter(str.isdigit, self.campo_telefone_paciente.value))[:11]
            fmt = ""
            if len(numeros) >= 1:
                fmt += "(" + numeros[:2] + ") "
            if len(numeros) >= 7:
                fmt += numeros[2:7] + "-"
            if len(numeros) > 7:
                fmt += numeros[7:]
            elif len(numeros) > 2:
                fmt += numeros[2:7]
            self.campo_telefone_paciente.value = fmt
            self.campo_telefone_paciente.update()

        def nascimento_change_cadastro(e):
            texto_original = self.campo_nascimento.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:8]

            if len(numeros) == 8:
                pass

        def nascimento_blur_cadastro(e):
            texto_original = self.campo_nascimento.value
            numeros = ''.join(filter(str.isdigit, texto_original))[:8]
            formatado = ""
            if len(numeros) >= 2:
                formatado += numeros[:2] + "/"
            if len(numeros) >= 4:
                formatado += numeros[2:4] + "/"
            if len(numeros) > 4:
                formatado += numeros[4:]
            elif len(numeros) > 2:
                formatado += numeros[2:]
            self.campo_nascimento.value = formatado
            self.campo_nascimento.update()

        self.current_view.controls.clear()

        # Campos do formulário
        self.campo_nome_paciente = ft.TextField(label="Nome do Paciente", border_radius=10, bgcolor="#F3F4F6")
        self.campo_cpf = ft.TextField(label="CPF", on_blur=cpf_blur, on_change=cpf_change_cadastro, border_radius=10, bgcolor="#F3F4F6")
        self.campo_email = ft.TextField(label="Email", border_radius=10, bgcolor="#F3F4F6")
        self.campo_nascimento = ft.TextField(label="Data de Nascimento (DD/MM/AAAA)", on_blur=nascimento_blur_cadastro, on_change=nascimento_change_cadastro, border_radius=10, bgcolor="#F3F4F6")
        self.campo_telefone_paciente = ft.TextField(label="Telefone", on_blur=telefone_blur_cadastro, on_change=telefone_change_cadastro, border_radius=10, bgcolor="F3F4F6")
        self.campo_senha = ft.TextField(label="Senha", password=True, can_reveal_password=True, border_radius=10, bgcolor="#F3F4F6")
        self.campo_confirmar_senha = ft.TextField(label="Confirmar Senha", password=True, can_reveal_password=True, border_radius=10, bgcolor="#F3F4F6")

        # Estrutura do formulário
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column(
                    [
                        ft.Text("Cadastrar Novo Paciente", size=28, weight="bold", color=ft.Colors.BLUE_900),
                        ft.Container(height=20),
                        ft.Container(
                            padding=30,
                            bgcolor="#FFFFFF",
                            border_radius=20,
                            shadow=ft.BoxShadow(blur_radius=20, spread_radius=2, color=ft.Colors.BLACK26, offset=ft.Offset(0, 8)),
                            content=ft.Column(
                                [
                                    self.campo_nome_paciente,
                                    self.campo_cpf,
                                    self.campo_email,
                                    self.campo_nascimento,
                                    self.campo_telefone_paciente,
                                    self.campo_senha,
                                    self.campo_confirmar_senha,
                                    ft.Container(height=20),
                                    ft.Row(
                                        [
                                            ft.ElevatedButton(
                                                "Salvar Paciente",
                                                bgcolor=ft.Colors.BLUE_600,
                                                color="white",
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                expand=True,
                                                on_click=self.salvar_paciente
                                            ),
                                            ft.OutlinedButton(
                                                "Cancelar",
                                                expand=True,
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                on_click=lambda e: self.load_pacientes(),
                                            ),
                                        ],
                                        spacing=20,
                                    )
                                ],
                                spacing=15,
                            ),
                        )
                    ],
                    spacing=20,
                    expand=True,
                    alignment=ft.MainAxisAlignment.CENTER,
                ),
            )
        )

        self.page.update()

    def salvar_paciente(self, e):
        nome = self.campo_nome_paciente.value.strip()
        cpf = self.campo_cpf.value.strip()
        email = self.campo_email.value.strip()
        nascimento = self.campo_nascimento.value.strip()
        telefone = self.campo_telefone_paciente.value.strip()  # <-- NOVO
        senha = self.campo_senha.value.strip()
        confirmar_senha = self.campo_confirmar_senha.value.strip()

        # Validação dos campos
        if not nome or not cpf or not email or not nascimento or not telefone or not senha or not confirmar_senha:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Todos os campos são obrigatórios."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        
        if senha != confirmar_senha:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("As senhas não coincidem."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        # Formata a data de nascimento para o formato esperado no banco (DD/MM/AAAA para AAAA-MM-DD)
        try:
            dia, mes, ano = nascimento.split("/")
            nascimento_formatado = f"{ano}-{mes}-{dia}"
        except ValueError:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Data de nascimento inválida."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        # Tenta registrar o usuário no banco de dados
        sucesso = registrar_usuario(nome, email, cpf, nascimento_formatado, telefone, senha)
        
        if sucesso:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Paciente cadastrado com sucesso!"), bgcolor="green")
            self.page.snack_bar.open = True
            self.page.update()
            self.load_pacientes()  # Retorna para a lista de pacientes
        else:
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Erro: CPF ou email já cadastrado."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()

    def load_cadastro_agendamento(self, e=None):
        pacientes = listar_usuarios()
        medicamentos = listar_medicamentos()

        self.dropdown_paciente = ft.Dropdown(
            label="Paciente",
            options=[ft.dropdown.Option(str(u[0]), u[1]) for u in pacientes],
            border_radius=10,
            bgcolor="#F9FAFB",
            expand=True
        )

        self.dropdown_medicamento = ft.Dropdown(
            label="Medicamento",
            options=[ft.dropdown.Option(str(m[0]), m[1]) for m in medicamentos],
            border_radius=10,
            bgcolor="#F9FAFB",
            expand=True,
            on_change=self.atualizar_farmacia_automatica
        )

        self.texto_farmacia = ft.Text("Selecione um medicamento...", size=14, weight="bold", color=ft.Colors.BLUE_900)

        # DatePicker
        self.data_escolhida = None
        self.date_picker = ft.DatePicker(
            on_change=self.selecionar_data,
            first_date=datetime(2024, 1, 1),
            last_date=datetime(2026, 12, 31),
        )

        if self.date_picker not in self.page.overlay:
            self.page.overlay.append(self.date_picker)

        self.botao_calendario = ft.ElevatedButton(
            "📅 Escolher Data",
            on_click=lambda _: self.abrir_date_picker(),
            style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=10)),
        )

        self.texto_data = ft.Text("Nenhuma data selecionada", size=14, italic=True)

        # Substituição: Dropdown com horários fixos
        horas_disponiveis = [f"{h:02d}:{m:02d}" for h in range(8, 18) for m in (0, 30)]
        self.campo_horario = ft.Dropdown(
            label="⏰ Horário de Retirada",
            hint_text="Selecione o horário desejado",
            options=[ft.dropdown.Option(h) for h in horas_disponiveis],
            width=400,  # 👈 aumenta o botão
            border_radius=12,
            bgcolor=ft.Colors.WHITE,
            filled=True,
            focused_border_color=ft.Colors.BLUE_900,
            border_color="#CBD5E1",
            text_style=ft.TextStyle(size=14, weight="bold", color="#1E3A8A"),
            label_style=ft.TextStyle(size=14, weight="bold", color="#1E3A8A"),
            icon=ft.Icons.ACCESS_TIME
        )


        self.current_view.controls.clear()
        self.current_view.controls.append(
            ft.Container(
                padding=30,
                content=ft.Column(
                    [
                        ft.Text("Cadastrar Novo Agendamento", size=28, weight="bold", color=ft.Colors.BLUE_900),
                        ft.Container(height=20),
                        ft.Container(
                            padding=30,
                            bgcolor="#FFFFFF",
                            border_radius=20,
                            shadow=ft.BoxShadow(blur_radius=20, spread_radius=2, color="#CBD5E1", offset=ft.Offset(0, 8)),
                            content=ft.Column(
                                [
                                    self.dropdown_paciente,
                                    self.dropdown_medicamento,
                                    self.texto_farmacia,
                                    ft.Column([self.botao_calendario, self.texto_data]),
                                    self.campo_horario,
                                    ft.Container(height=20),
                                    ft.Row(
                                        [
                                            ft.ElevatedButton(
                                                "Salvar Agendamento",
                                                bgcolor=ft.Colors.BLUE_600,
                                                color="white",
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                expand=True,
                                                on_click=self.salvar_agendamento
                                            ),
                                            ft.OutlinedButton(
                                                "Cancelar",
                                                expand=True,
                                                height=50,
                                                style=ft.ButtonStyle(shape=ft.RoundedRectangleBorder(radius=12)),
                                                on_click=lambda e: self.load_agendamentos(),
                                            ),
                                        ],
                                        spacing=20,
                                    )
                                ],
                                spacing=15,
                            ),
                        )
                    ],
                    spacing=20,
                    expand=True,
                    alignment=ft.MainAxisAlignment.CENTER,
                ),
            )
        )
        self.page.update()

    def abrir_date_picker(self):
        self.date_picker.open = True
        self.page.update()

    def filtrar_agendamentos(self, e):
        termo = self.campo_busca_agendamentos.value.strip().lower()
        agendamentos = listar_agendamentos()

        resultado = [
            a for a in agendamentos
            if termo in a[1].lower() or termo in a[2].lower() or termo in a[3].lower()
        ]

        self.tabela_agendamentos_ref.current.rows = self.gerar_rows_agendamentos(resultado)
        self.tabela_agendamentos_ref.current.update()


    def selecionar_data(self, e):
        if self.date_picker.value:
            self.data_escolhida = self.date_picker.value
            self.texto_data.value = f"Data selecionada: {self.data_escolhida.strftime('%d/%m/%Y')}"
            self.texto_data.update()


    def salvar_agendamento(self, e=None):
        from datetime import datetime
        paciente_id = self.dropdown_paciente.value
        medicamento_id = self.dropdown_medicamento.value
        farmacia_id = self.farmacia_automatica["id"] if hasattr(self, "farmacia_automatica") else None
        codigo_medicamento = "AG" + datetime.now().strftime("%Y%m%d%H%M%S")
        if self.data_escolhida:
            data = self.data_escolhida.strftime('%Y-%m-%d')
        else:
            data = ""
        horario = self.campo_horario.value.strip()

        if not all([paciente_id, medicamento_id, farmacia_id, data, horario]):
            self.page.snack_bar = ft.SnackBar(content=ft.Text("Todos os campos são obrigatórios."), bgcolor="red")
            self.page.snack_bar.open = True
            self.page.update()
            return

        adicionar_agendamento(int(paciente_id), int(medicamento_id), int(farmacia_id), codigo_medicamento, data, horario, status="Pendente")

        self.page.snack_bar = ft.SnackBar(content=ft.Text("Agendamento cadastrado com sucesso."), bgcolor="green")
        self.page.snack_bar.open = True
        self.page.update()
        self.load_agendamentos()


    def atualizar_farmacia_automatica(self, e):
        med_id = int(self.dropdown_medicamento.value)
        med_info = next((m for m in listar_medicamentos() if m[0] == med_id), None)

        if med_info:
            self.farmacia_automatica = {
                "id": med_info[10],
                "nome": med_info[8],
                "endereco": med_info[9],
            }
            self.texto_farmacia.value = f"🏥 {self.farmacia_automatica['nome']}\n📍 {self.farmacia_automatica['endereco']}"
            self.texto_farmacia.update()

    def gerar_relatorio_pdf(self):
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor, lightgrey, black
        from datetime import datetime
        from collections import Counter, defaultdict
        import calendar

        caminho = "relatorio_farmconnect.pdf"
        c = canvas.Canvas(caminho, pagesize=A4)
        largura, altura = A4
        y = altura - 2 * cm

        def nova_pagina():
            nonlocal y
            c.showPage()
            y = altura - 2 * cm

        def header(titulo, subtitulo=""):
            nonlocal y
            c.setFont("Helvetica-Bold", 16)
            c.setFillColor(HexColor("#1E3A8A"))
            c.drawString(2 * cm, y, titulo)
            y -= 0.6 * cm
            if subtitulo:
                c.setFont("Helvetica", 10)
                c.setFillColor(HexColor("#374151"))
                c.drawString(2 * cm, y, subtitulo)
                y -= 0.5 * cm
            c.setStrokeColor(HexColor("#1E3A8A"))
            c.line(2 * cm, y, largura - 2 * cm, y)
            y -= 0.6 * cm
            c.setFillColor(black)

        def rodape():
            c.setFont("Helvetica-Oblique", 9)
            c.setFillColor(HexColor("#6B7280"))
            c.drawString(2 * cm, 1.5 * cm, f"Relatório gerado automaticamente em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
            c.setFillColor(black)

        # PACIENTES
        pacientes = listar_usuarios()
        header("📋 Pacientes Cadastrados", f"Total: {len(pacientes)}")
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2 * cm, y, "ID")
        c.drawString(3.5 * cm, y, "Nome")
        c.drawString(9 * cm, y, "Email")
        c.drawString(14.5 * cm, y, "Status")
        y -= 0.4 * cm
        c.line(2 * cm, y, largura - 2 * cm, y)
        y -= 0.4 * cm
        c.setFont("Helvetica", 9)
        pacientes_mes = Counter()
        status_pacientes = defaultdict(Counter)
        for u in pacientes:
            if y < 3 * cm:
                nova_pagina()
                c.setFont("Helvetica-Bold", 10)
                c.drawString(2 * cm, y, "ID")
                c.drawString(3.5 * cm, y, "Nome")
                c.drawString(9 * cm, y, "Email")
                c.drawString(14.5 * cm, y, "Status")
                y -= 0.4 * cm
                c.line(2 * cm, y, largura - 2 * cm, y)
                y -= 0.4 * cm
                c.setFont("Helvetica", 9)
            c.drawString(2 * cm, y, str(u[0]))
            c.drawString(3.5 * cm, y, u[1][:30])
            c.drawString(9 * cm, y, u[2][:35])
            c.drawString(14.5 * cm, y, u[7])
            y -= 0.4 * cm
            try:
                mes = calendar.month_name[datetime.strptime(u[6], "%Y-%m-%d %H:%M:%S").month]
                pacientes_mes[mes] += 1
                status_pacientes[mes][u[7]] += 1
            except:
                pass

        # AGENDAMENTOS
        nova_pagina()
        agendamentos = listar_agendamentos()
        header("📅 Agendamentos Realizados", f"Total: {len(agendamentos)}")
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2 * cm, y, "ID")
        c.drawString(3.5 * cm, y, "Paciente")
        c.drawString(8 * cm, y, "Medicamento")
        c.drawString(12.5 * cm, y, "Data")
        c.drawString(15 * cm, y, "Hora")
        c.drawString(17 * cm, y, "Status")
        y -= 0.4 * cm
        c.line(2 * cm, y, largura - 2 * cm, y)
        y -= 0.4 * cm
        c.setFont("Helvetica", 9)
        agendamentos_mes = Counter()
        status_agendamento = defaultdict(Counter)
        medicamentos_mes = defaultdict(Counter)
        farmacias_mes = defaultdict(Counter)
        for a in agendamentos:
            if y < 3 * cm:
                nova_pagina()
                c.setFont("Helvetica-Bold", 10)
                c.drawString(2 * cm, y, "ID")
                c.drawString(3.5 * cm, y, "Paciente")
                c.drawString(8 * cm, y, "Medicamento")
                c.drawString(12.5 * cm, y, "Data")
                c.drawString(15 * cm, y, "Hora")
                c.drawString(17 * cm, y, "Status")
                y -= 0.4 * cm
                c.line(2 * cm, y, largura - 2 * cm, y)
                y -= 0.4 * cm
                c.setFont("Helvetica", 9)
            c.drawString(2 * cm, y, str(a[0]))
            c.drawString(3.5 * cm, y, a[1][:25])
            c.drawString(8 * cm, y, a[2][:25])
            c.drawString(12.5 * cm, y, a[5])
            c.drawString(15 * cm, y, a[6])
            c.drawString(17 * cm, y, a[7])
            y -= 0.4 * cm
            try:
                # usa data_criacao do agendamento (a[8]) para o agrupamento mensal
                mes = calendar.month_name[datetime.strptime(a[8], "%Y-%m-%d %H:%M:%S").month]
                agendamentos_mes[mes] += 1
                status_agendamento[mes][a[7]] += 1
                medicamentos_mes[mes][a[2]] += 1
                farmacias_mes[mes][a[3]] += 1
            except:
                pass

        # REAGENDAMENTOS
        nova_pagina()
        reagendamentos = listar_reagendamentos()
        header("🔁 Reagendamentos", f"Total: {len(reagendamentos)}")

        c.setFont("Helvetica-Bold", 10)
        c.drawString(2 * cm, y, "ID")
        c.drawString(3.5 * cm, y, "Paciente")
        c.drawString(8 * cm, y, "Medicamento")
        c.drawString(12.5 * cm, y, "Antigo (Data Hora)")
        c.drawString(16.5 * cm, y, "Novo (Data Hora)")
        y -= 0.4 * cm
        c.line(2 * cm, y, largura - 2 * cm, y)
        y -= 0.4 * cm
        c.setFont("Helvetica", 9)

        reag_mes = Counter()  # contagem mensal de reagendamentos pelo 'criado_em'
        for r in reagendamentos:
            if y < 3 * cm:
                nova_pagina()
                c.setFont("Helvetica-Bold", 10)
                c.drawString(2 * cm, y, "ID")
                c.drawString(3.5 * cm, y, "Paciente")
                c.drawString(8 * cm, y, "Medicamento")
                c.drawString(12.5 * cm, y, "Antigo (Data Hora)")
                c.drawString(16.5 * cm, y, "Novo (Data Hora)")
                y -= 0.4 * cm
                c.line(2 * cm, y, largura - 2 * cm, y)
                y -= 0.4 * cm
                c.setFont("Helvetica", 9)

            antigo = f"{r[5]} {r[6]}" if r[5] and r[6] else "-"
            novo   = f"{r[7]} {r[8]}" if r[7] and r[8] else "-"

            c.drawString(2 * cm, y, str(r[0]))      # ID
            c.drawString(3.5 * cm, y, r[1][:25])    # Paciente
            c.drawString(8 * cm, y, r[2][:25])      # Medicamento
            c.drawString(12.5 * cm, y, antigo[:22])
            c.drawString(16.5 * cm, y, novo[:22])
            y -= 0.4 * cm

            # contabiliza mês do criado_em (r[9])
            try:
                mes_r = calendar.month_name[datetime.strptime(r[9], "%Y-%m-%d %H:%M:%S").month]
                reag_mes[mes_r] += 1
            except:
                pass

        # COMPARATIVO MENSAL
        nova_pagina()
        header("📊 Comparativo Mensal")
        # inclui meses que apareceram em pacientes, agendamentos ou reagendamentos
        meses_todos = sorted(
            set(pacientes_mes.keys()) | set(agendamentos_mes.keys()) | set(reag_mes.keys()),
            key=lambda m: list(calendar.month_name).index(m)
        )

        c.setFont("Helvetica-Bold", 10)
        c.drawString(2 * cm, y, "Mês")
        c.drawString(6 * cm, y, "Pacientes")
        c.drawString(9 * cm, y, "Medicamentos")
        c.drawString(13 * cm, y, "Agendamentos")
        c.drawString(17 * cm, y, "Reagend.")
        y -= 0.4 * cm
        c.line(2 * cm, y, largura - 2 * cm, y)
        y -= 0.4 * cm
        c.setFont("Helvetica", 9)

        for mes in meses_todos:
            if y < 3 * cm:
                nova_pagina()
                c.setFont("Helvetica-Bold", 10)
                c.drawString(2 * cm, y, "Mês")
                c.drawString(6 * cm, y, "Pacientes")
                c.drawString(9 * cm, y, "Medicamentos")
                c.drawString(13 * cm, y, "Agendamentos")
                c.drawString(17 * cm, y, "Reagend.")
                y -= 0.4 * cm
                c.line(2 * cm, y, largura - 2 * cm, y)
                y -= 0.4 * cm
                c.setFont("Helvetica", 9)

            total_pac = pacientes_mes.get(mes, 0)
            total_ag = agendamentos_mes.get(mes, 0)
            total_med = sum(medicamentos_mes[mes].values()) if mes in medicamentos_mes else 0
            total_reag = reag_mes.get(mes, 0)

            c.drawString(2 * cm, y, mes[:10])
            c.drawString(6 * cm, y, str(total_pac))
            c.drawString(9 * cm, y, str(total_med))
            c.drawString(13 * cm, y, str(total_ag))
            c.drawString(17 * cm, y, str(total_reag))
            y -= 0.4 * cm

        rodape()
        c.save()
        self.page.launch_url(caminho)

    
    
    def gerar_relatorio_excel(self):
        wb = openpyxl.Workbook()
        del wb["Sheet"]

        # ===== estilos =====
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        align_left = Alignment(horizontal="left", vertical="center")
        title_font = Font(size=14, bold=True, color="FFFFFF")
        border_preta = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000")
        )

        def autoajustar_colunas(ws):
            for column_cells in ws.columns:
                max_length = max(len(str(cell.value or "")) for cell in column_cells)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = max(max_length + 2, 12)

        def formatar_titulo(ws, texto, total_cols):
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
            cell = ws.cell(row=1, column=1)
            cell.value = texto
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

        def escrever_cabecalho(ws, headers, linha=2):
            for i, cab in enumerate(headers, start=1):
                cell = ws.cell(row=linha, column=i, value=cab)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_left
                cell.border = border_preta

        def formatar_corpo(ws, linha_inicial):
            for row in ws.iter_rows(min_row=linha_inicial, max_row=ws.max_row):
                for cell in row:
                    cell.alignment = align_left
                    cell.border = border_preta

        # Pacientes
        pacientes = listar_usuarios()
        ws1 = wb.create_sheet("Pacientes")
        headers_pac = ["ID", "Nome", "Email", "CPF", "Nascimento", "Telefone", "Criado em", "Status"]
        formatar_titulo(ws1, "Pacientes Cadastrados", len(headers_pac))
        escrever_cabecalho(ws1, headers_pac)
        for u in pacientes:
            ws1.append([u[0], u[1], u[2], u[3], u[4], u[5], u[6], u[7]])
        formatar_corpo(ws1, linha_inicial=3)
        autoajustar_colunas(ws1)

        # Agendamentos
        agendamentos = listar_agendamentos()
        ws2 = wb.create_sheet("Agendamentos")
        headers_ag = ["ID", "Paciente", "Medicamento", "Farmácia", "Código", "Data", "Horário", "Status", "Criado em"]
        formatar_titulo(ws2, "Agendamentos Realizados", len(headers_ag))
        escrever_cabecalho(ws2, headers_ag)
        for a in agendamentos:
            ws2.append([a[0], a[1], a[2], a[3], a[4], a[5], a[6], a[7], a[8]])
        formatar_corpo(ws2, linha_inicial=3)
        autoajustar_colunas(ws2)

        # Reagendamentos (sem total)
        def listar_reagendamentos_excel():
            conn = sqlite3.connect("farmconnect.db")
            cur = conn.cursor()
            cur.execute("""
                SELECT
                    r.id,
                    u.nome           AS paciente,
                    m.nome           AS medicamento,
                    f.nome           AS farmacia,
                    a.codigo         AS codigo_agendamento,
                    r.data_antiga,
                    r.horario_antigo,
                    r.data_nova,
                    r.horario_novo,
                    a.status         AS status_atual,
                    COALESCE(r.criado_em, a.data_criacao) AS criado_em
                FROM reagendamentos r
                JOIN agendamentos a ON a.id = r.agendamento_id
                JOIN usuarios    u ON u.id = r.usuario_id
                JOIN medicamentos m ON m.id = a.medicamento_id
                JOIN farmacias    f ON f.id = a.farmacia_id
                ORDER BY r.id DESC
            """)
            rows = cur.fetchall()
            cur.close()
            conn.close()
            return rows

        reag = listar_reagendamentos_excel()
        ws4 = wb.create_sheet("Reagendamentos")
        headers_reag = [
            "ID", "Paciente", "Medicamento", "Farmácia", "Código",
            "Data antiga", "Hora antiga", "Data nova", "Hora nova",
            "Status atual", "Criado em"
        ]
        formatar_titulo(ws4, "Reagendamentos", len(headers_reag))
        escrever_cabecalho(ws4, headers_reag)
        for r in reag:
            ws4.append([r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7], r[8], r[9], r[10]])
        formatar_corpo(ws4, linha_inicial=3)
        autoajustar_colunas(ws4)

        # Medicamentos
        medicamentos = listar_medicamentos()
        ws3 = wb.create_sheet("Medicamentos")
        headers_med = ["ID", "Nome", "Código", "Descrição", "Estoque", "Categoria", "Fabricante", "Farmácia", "Ativo"]
        formatar_titulo(ws3, "Medicamentos Cadastrados", len(headers_med))
        escrever_cabecalho(ws3, headers_med)
        for m in medicamentos:
            ws3.append([
                m[0], m[1], m[2], m[3], m[5] or 0, m[6] or "-", m[7] or "-", m[8] or "-", "Sim" if m[9] else "Não"
            ])
        formatar_corpo(ws3, linha_inicial=3)
        autoajustar_colunas(ws3)

        # Salvar
        nome_arquivo = f"relatorio_farmconnect_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        wb.save(nome_arquivo)
        self.page.launch_url(nome_arquivo)


    def build_tela(self):
        return ft.View(
            route="/admin_dashboard",
            padding=0,
            controls=[
                ft.Row(
                    expand=True,
                    spacing=0,
                    vertical_alignment=ft.CrossAxisAlignment.START,
                    controls=[
                        self.side_menu(),
                        ft.Column(
                            expand=True,
                            spacing=0,
                            controls=[
                                self.header(),
                                ft.Container(
                                    expand=True,
                                    padding=20,
                                    content=ft.Column(
                                        expand=True,
                                        scroll=ft.ScrollMode.ADAPTIVE,
                                        controls=[
                                            self.current_view
                                        ]
                                    )
                                )
                            ]
                        )
                    ]
                )
            ]
        )

if __name__ == "__main__":
    def main(page: ft.Page):
        app = TelaAdminDashboard(page)
        page.views.append(app.build_tela()) # app.build_tela()
        page.update()

    ft.app(target=main)  # Ensure the main function is passed as the target
